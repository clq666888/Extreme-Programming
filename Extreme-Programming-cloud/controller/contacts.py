from flask import Blueprint, request, jsonify, send_file
import database
import sqlite3
import openpyxl
from io import BytesIO

contacts_bp = Blueprint('contacts', __name__)

@contacts_bp.route("/", methods=["GET"])
def get_contacts():
    search = request.args.get('search', '').strip()
    favorite = request.args.get('favorite', '').lower() == 'true'
    conn = database.get_connection()
    c = conn.cursor()

    query = "SELECT * FROM contacts"
    params = []
    if search:
        query += " WHERE name LIKE ? OR id IN (SELECT contact_id FROM contact_details WHERE value LIKE ?)"
        params += [f"%{search}%", f"%{search}%"]
    elif favorite:
        query += " WHERE is_favorite=1"

    c.execute(query, params)
    contacts = [dict(row) for row in c.fetchall()]

    for contact in contacts:
        c.execute("SELECT type, value FROM contact_details WHERE contact_id=?", (contact['id'],))
        contact['details'] = [dict(d) for d in c.fetchall()]

    conn.close()
    return jsonify(contacts)

@contacts_bp.route("/", methods=["POST"])
def add_contact():
    data = request.get_json()
    name = data['name']
    details = data.get('details', [])
    conn = database.get_connection()
    c = conn.cursor()
    c.execute("INSERT INTO contacts (name) VALUES (?)", (name,))
    contact_id = c.lastrowid
    for d in details:
        c.execute("INSERT INTO contact_details (contact_id, type, value) VALUES (?, ?, ?)",
                  (contact_id, d['type'], d['value']))
    conn.commit()
    conn.close()
    return jsonify({"id": contact_id})

@contacts_bp.route("/<int:contact_id>", methods=["PUT"])
def edit_contact(contact_id):
    data = request.get_json()
    name = data['name']
    details = data.get('details', [])
    conn = database.get_connection()
    c = conn.cursor()
    c.execute("UPDATE contacts SET name=? WHERE id=?", (name, contact_id))
    c.execute("DELETE FROM contact_details WHERE contact_id=?", (contact_id,))
    for d in details:
        c.execute("INSERT INTO contact_details (contact_id, type, value) VALUES (?, ?, ?)",
                  (contact_id, d['type'], d['value']))
    conn.commit()
    conn.close()
    return jsonify({"id": contact_id})

@contacts_bp.route("/<int:contact_id>", methods=["DELETE"])
def delete_contact(contact_id):
    conn = database.get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM contacts WHERE id=?", (contact_id,))
    conn.commit()
    conn.close()
    return jsonify({"deleted": contact_id})

@contacts_bp.route("/<int:contact_id>/favorite", methods=["PUT"])
def toggle_favorite(contact_id):
    data = request.get_json()
    is_fav = data.get('is_favorite', False)
    conn = database.get_connection()
    c = conn.cursor()
    c.execute("UPDATE contacts SET is_favorite=? WHERE id=?", (1 if is_fav else 0, contact_id))
    conn.commit()
    conn.close()
    return jsonify({"id": contact_id, "is_favorite": is_fav})

# ===== 导出 Excel =====
@contacts_bp.route("/export", methods=["GET"])
def export_contacts():
    conn = database.get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM contacts")
    contacts = [dict(row) for row in c.fetchall()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # 表头
    headers = ["姓名", "收藏", "电话", "邮箱", "QQ", "微信", "地址"]
    ws.append(headers)

    for contact in contacts:
        c.execute("SELECT type, value FROM contact_details WHERE contact_id=?", (contact["id"],))
        details = {row["type"]: row["value"] for row in c.fetchall()}

        row = [
            contact["name"],
            "是" if contact.get("is_favorite") else "否",
            details.get("phone", ""),
            details.get("email", ""),
            details.get("qq", ""),
            details.get("weixin", ""),
            details.get("address", "")
        ]
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    conn.close()

    return send_file(
        output,
        as_attachment=True,
        download_name="contacts.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ===== 导入 Excel =====
@contacts_bp.route("/import", methods=["POST"])
def import_contacts():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    conn = database.get_connection()
    c = conn.cursor()

    # 从第二行开始读取
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue  # 跳过空行
        name = row[0]
        is_fav = 1 if row[1] == "是" else 0
        details = {
            "phone": row[2] or "",
            "email": row[3] or "",
            "qq": row[4] or "",
            "weixin": row[5] or "",
            "address": row[6] or ""
        }

        c.execute("INSERT INTO contacts (name, is_favorite) VALUES (?, ?)", (name, is_fav))
        contact_id = c.lastrowid

        for t, v in details.items():
            if v:
                c.execute(
                    "INSERT INTO contact_details (contact_id, type, value) VALUES (?, ?, ?)",
                    (contact_id, t, v)
                )

    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})

