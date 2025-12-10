from flask import Blueprint, request, jsonify, send_file
import database
import sqlite3
import openpyxl
from io import BytesIO
import json

contacts_bp = Blueprint("contacts_bp", __name__)

def get_conn():
    return database.get_connection()

# 获取联系人列表
@contacts_bp.route("", methods=["GET"])
def get_contacts():
    search = request.args.get("search", "")
    favorite = request.args.get("favorite")
    conn = get_conn()
    c = conn.cursor()
    query = "SELECT * FROM contacts"
    params = []
    if search or favorite:
        query += " WHERE "
        conditions = []
        if search:
            conditions.append("name LIKE ? OR id IN (SELECT contact_id FROM contact_details WHERE value LIKE ?)")
            params += [f"%{search}%", f"%{search}%"]
        if favorite == "true":
            conditions.append("is_favorite=1")
        query += " AND ".join(conditions)
    c.execute(query, params)
    contacts = [dict(row) for row in c.fetchall()]

    # 加载详细信息
    for contact in contacts:
        c.execute("SELECT type, value FROM contact_details WHERE contact_id=?", (contact["id"],))
        contact["details"] = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(contacts)

# 新建联系人
@contacts_bp.route("", methods=["POST"])
def add_contact():
    data = request.json
    name = data.get("name")
    details = data.get("details", [])
    conn = get_conn()
    c = conn.cursor()
    c.execute("INSERT INTO contacts (name) VALUES (?)", (name,))
    contact_id = c.lastrowid
    for d in details:
        c.execute("INSERT INTO contact_details (contact_id,type,value) VALUES (?,?,?)",
                  (contact_id, d["type"], d["value"]))
    conn.commit()
    conn.close()
    return jsonify({"id": contact_id}), 201

# 更新联系人
@contacts_bp.route("/<int:id>", methods=["PUT"])
def update_contact(id):
    data = request.json
    name = data.get("name")
    details = data.get("details", [])
    conn = get_conn()
    c = conn.cursor()
    c.execute("UPDATE contacts SET name=? WHERE id=?", (name, id))
    c.execute("DELETE FROM contact_details WHERE contact_id=?", (id,))
    for d in details:
        c.execute("INSERT INTO contact_details (contact_id,type,value) VALUES (?,?,?)",
                  (id, d["type"], d["value"]))
    conn.commit()
    conn.close()
    return jsonify({"id": id})

# 删除联系人
@contacts_bp.route("/<int:id>", methods=["DELETE"])
def delete_contact(id):
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM contacts WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return jsonify({"deleted": True})

# 收藏/取消收藏
@contacts_bp.route("/<int:id>/favorite", methods=["PUT"])
def favorite_contact(id):
    data = request.json
    is_fav = data.get("is_favorite", False)
    conn = get_conn()
    c = conn.cursor()
    c.execute("UPDATE contacts SET is_favorite=? WHERE id=?", (1 if is_fav else 0, id))
    conn.commit()
    conn.close()
    return jsonify({"id": id, "is_favorite": is_fav})

# ---------------- 导出 Excel ----------------
@contacts_bp.route("/export", methods=["GET"])
def export_contacts():
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM contacts")
    contacts = [dict(row) for row in c.fetchall()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # 表头
    headers = ["姓名", "收藏", "电话", "邮箱", "QQ", "微信", "地址"]
    ws.append(headers)

    for contact in contacts:
        c.execute("SELECT type,value FROM contact_details WHERE contact_id=?", (contact["id"],))
        details = {row["type"]: row["value"] for row in c.fetchall()}
        row = [
            contact["name"],
            "是" if contact["is_favorite"] else "否",
            details.get("phone", ""),
            details.get("email", ""),
            details.get("qq", ""),
            details.get("weixin", ""),
            details.get("address", "")
        ]
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    conn.close()

    return send_file(output, as_attachment=True, download_name="contacts.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------- 导入 Excel ----------------
@contacts_bp.route("/import", methods=["POST"])
def import_contacts():
    if "file" not in request.files:
        return jsonify({"error":"No file"}), 400
    file = request.files["file"]
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    conn = get_conn()
    c = conn.cursor()

    # 从第二行开始读取
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        is_fav = 1 if row[1] == "是" else 0
        details = {
            "phone": row[2] or "",
            "email": row[3] or "",
            "qq": row[4] or "",
            "weixin": row[5] or "",
            "address": row[6] or ""
        }
        c.execute("INSERT INTO contacts (name,is_favorite) VALUES (?,?)", (name,is_fav))
        contact_id = c.lastrowid
        for t,v in details.items():
            if v:
                c.execute("INSERT INTO contact_details (contact_id,type,value) VALUES (?,?,?)", (contact_id,t,v))

    conn.commit()
    conn.close()
    return jsonify({"imported": True})
